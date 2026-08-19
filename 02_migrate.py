# -*- coding: utf-8 -*-
"""
STOCK RADAR · 기존 데이터 → Supabase 마이그레이션
=================================================

사용법
------
    pip install psycopg2-binary openpyxl pandas

    export SUPABASE_DB_URL="postgresql://postgres.xxxx:비밀번호@aws-0-ap-northeast-2.pooler.supabase.com:5432/postgres"

    python 02_migrate.py stocks     # ① 종목 마스터 (먼저 실행 — FK 대상)
    python 02_migrate.py price      # ② 전종목시세.xlsx     → daily_price
    python 02_migrate.py flow       # ③ 수급주체정리.xlsx   → daily_flow
    python 02_migrate.py kiwoom     # ④ 키움 CSV            → kiwoom_holder_stats
    python 02_migrate.py all        # 전부 순서대로
    python 02_migrate.py check      # 적재 결과 점검

접속 문자열 얻는 곳
-----------------
    Supabase 대시보드 → Project Settings → Database → Connection string
    ⚠ "Direct connection"은 IPv6 전용이라 GitHub Actions·Colab에서 안 붙는 경우가 많습니다.
       반드시 **Session pooler** 또는 **Transaction pooler** 문자열을 쓰세요.
       (호스트에 `pooler.supabase.com` 이 들어간 것)

설계 메모
--------
· 전부 UPSERT(ON CONFLICT DO UPDATE)라 몇 번을 다시 돌려도 중복이 생기지 않습니다.
· 87MB 엑셀은 openpyxl read_only 스트리밍으로 읽습니다(통째 로드 시 수 분 + 대용량 메모리).
· 시트마다 헤더 행이 다릅니다(외국인 4행 / 기타법인 1행 / 사모 12행) — 자동 탐색합니다.
· 외국인·기관합계 시트는 "시총 5000억↑·거래량 1억↑" 필터로 수집돼 공백이 있습니다.
  해당 행에 is_partial=true를 세워 신호 계산에서 제외할 수 있게 합니다.
"""

import os
import re
import sys
import csv
import time

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

DB_URL      = os.getenv("SUPABASE_DB_URL", "")
XL_PRICE    = os.getenv("XL_PRICE",  "전종목시세.xlsx")
XL_FLOW     = os.getenv("XL_FLOW",   "수급주체정리.xlsx")
CSV_KIWOOM  = os.getenv("CSV_KIWOOM", "20260814_키운보유종목통계.csv")

BATCH = 5000

# 수급 시트명 → daily_flow 컬럼
FLOW_SHEETS = {
    "외국인":    "foreign_net",
    "기관합계":  "inst_net",
    "금융투자":  "fin_inv_net",
    "투신":      "inv_trust_net",
    "사모":      "pe_net",
    "연기금 등": "pension_net",
    "기타법인":  "corp_other_net",
}
# 필터 수집돼 공백이 있는 시트
PARTIAL_SHEETS = {"외국인", "기관합계"}

ETF_PAT = re.compile(
    r"TIGER|KODEX|PLUS |ACE |RISE |SOL |HANARO|KOSEF|ARIRANG|TIMEFOLIO|KIWOOM |"
    r"마이다스|파워|삼성레버리지|미래에셋|한국투자|스팩|제[0-9]+호"
)


# ──────────────────────────────────────────────────────────────
# 공통
# ──────────────────────────────────────────────────────────────
def conn():
    if not DB_URL:
        sys.exit("❌ SUPABASE_DB_URL 환경변수를 설정하세요.")
    c = psycopg2.connect(DB_URL)
    c.autocommit = False
    return c


def norm_code(v):
    if v is None:
        return None
    s = str(v).strip().replace("'", "")
    s = s.split(".")[0]
    return s.zfill(6) if s.isdigit() else (s.upper() if s else None)


def to_date(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    s = str(v).strip()[:10].replace("/", "-").replace(".", "-")
    return s if re.match(r"^\d{4}-\d{2}-\d{2}$", s) else None


def to_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).replace(",", "").replace("+", "").strip()
    if s in ("", "-", "N/A"):
        return None
    try:
        return float(s) if "." in s else int(s)
    except ValueError:
        return None


def find_header(ws, need, max_scan=40):
    """
    헤더 행 자동 탐색 → (행번호, {컬럼명: 인덱스})

    실측 헤더 위치: 전종목시세 24행 / 수급 외국인·기관합계·금융투자·투신·연기금 4행
                    기타법인 1행 / 사모 12행.  넉넉히 40행까지 훑습니다.
    """
    for rno, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True), 1):
        vals = [str(v).strip() if v is not None else "" for v in row]
        if all(c in vals for c in need):
            return rno, {c: vals.index(c) for c in vals if c}
    return None, None


def dedup(rows, keylen, label=""):
    """
    (trade_date, code) 같은 복합키 중복 제거. 마지막 행이 이깁니다.

    필요한 이유: Postgres의 ON CONFLICT DO UPDATE는 **같은 INSERT 명령 안에**
    중복 키가 있으면 CardinalityViolation으로 실패합니다.
    실제 전종목시세.xlsx에도 2025-12-15 휴젤(145020)이 2번 들어 있었습니다
    (엑셀 병합 과정의 중복 append로 추정, 두 행의 값은 동일).
    """
    seen = {}
    for r in rows:
        seen[r[:keylen]] = r
    n = len(rows) - len(seen)
    if n:
        print(f"  ⚠ 중복 키 {n:,}건 제거{(' — ' + label) if label else ''}")
    return list(seen.values())


def push(cur, sql, rows, label):
    """배치 단위 upsert + 진행 표시"""
    done = 0
    for i in range(0, len(rows), BATCH):
        chunk = rows[i:i + BATCH]
        execute_values(cur, sql, chunk, page_size=BATCH)
        done += len(chunk)
        print(f"\r  {label}: {done:,}/{len(rows):,}", end="", flush=True)
    print()


def log(cur, job, status, n, msg=""):
    cur.execute(
        "insert into ingest_log(job,status,row_count,message) values (%s,%s,%s,%s)",
        (job, status, n, msg[:500]),
    )


# ──────────────────────────────────────────────────────────────
# ① 종목 마스터
# ──────────────────────────────────────────────────────────────
def migrate_stocks():
    print("\n[1/4] 종목 마스터")
    master = {}

    # (a) 수급주체정리.xlsx 업종 시트 — 종목코드·종목명·시장구분·업종명 (2,774종목)
    wb = openpyxl.load_workbook(XL_FLOW, read_only=True, data_only=True)
    if "업종" in wb.sheetnames:
        ws = wb["업종"]
        for row in ws.iter_rows(min_row=3, values_only=True):
            code = norm_code(row[1] if len(row) > 1 else None)
            if not code or not row[2]:
                continue
            master[code] = {
                "name": str(row[2]).strip(),
                "market": str(row[3]).strip() if len(row) > 3 and row[3] else None,
                "sector_krx": str(row[4]).strip() if len(row) > 4 and row[4] else None,
            }
        print(f"  업종 시트   : {len(master):,}종목 (업종 포함)")
    wb.close()

    # (b) 전종목시세.xlsx — 시장구분 보강
    try:
        wb = openpyxl.load_workbook(XL_PRICE, read_only=True, data_only=True)
        ws = wb["전종목시세"]
        hdr, idx = find_header(ws, ["날짜", "종목코드", "종목명", "종가"])
        if idx:
            n0 = len(master)
            for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
                code = norm_code(row[idx["종목코드"]] if idx["종목코드"] < len(row) else None)
                if not code:
                    continue
                nm = row[idx["종목명"]] if idx["종목명"] < len(row) else None
                mk = row[idx["시장구분"]] if "시장구분" in idx and idx["시장구분"] < len(row) else None
                e = master.setdefault(code, {"name": None, "market": None, "sector_krx": None})
                e["name"] = e["name"] or (str(nm).strip() if nm else None)
                e["market"] = e["market"] or (str(mk).strip() if mk else None)
            print(f"  전종목시세  : +{len(master)-n0:,}종목")
        wb.close()
    except FileNotFoundError:
        print(f"  ⚠ {XL_PRICE} 없음 — 건너뜀")

    # (c) 키움 CSV — ETF/스팩까지 포함한 최대 커버리지
    try:
        with open(CSV_KIWOOM, encoding="cp949") as f:
            n0 = len(master)
            for r in csv.DictReader(f):
                code = norm_code(r.get("종목코드"))
                if not code:
                    continue
                e = master.setdefault(code, {"name": None, "market": None, "sector_krx": None})
                e["name"] = e["name"] or (r.get("종목명") or "").strip()
            print(f"  키움 CSV    : +{len(master)-n0:,}종목")
    except FileNotFoundError:
        print(f"  ⚠ {CSV_KIWOOM} 없음 — 건너뜀")

    rows = []
    for code, e in master.items():
        name = e["name"] or code
        stype = "ETF" if ETF_PAT.search(name) else ("PREF" if name.endswith("우") else "STOCK")
        rows.append((code, name, e["market"], stype, e["sector_krx"]))

    stock_n = sum(1 for r in rows if r[3] == "STOCK")
    print(f"  → 합계 {len(rows):,}종목 (STOCK {stock_n:,} / 기타 {len(rows)-stock_n:,})")

    with conn() as c, c.cursor() as cur:
        push(cur, """
            insert into stocks (code,name,market,security_type,sector_krx)
            values %s
            on conflict (code) do update set
              name          = excluded.name,
              market        = coalesce(excluded.market, stocks.market),
              security_type = excluded.security_type,
              sector_krx    = coalesce(excluded.sector_krx, stocks.sector_krx),
              updated_at    = now()
        """, rows, "적재")
        log(cur, "master", "SUCCESS", len(rows), "excel+kiwoom seed")
        c.commit()
    print("  ✅ 완료")


# ──────────────────────────────────────────────────────────────
# ② 전종목시세 → daily_price
# ──────────────────────────────────────────────────────────────
def migrate_price():
    print("\n[2/4] 전종목시세 → daily_price")
    wb = openpyxl.load_workbook(XL_PRICE, read_only=True, data_only=True)
    ws = wb["전종목시세"]
    hdr, idx = find_header(ws, ["날짜", "종목코드", "종가", "거래량"])
    if not idx:
        sys.exit("❌ 헤더 행을 찾지 못했습니다.")
    print(f"  헤더 {hdr}행")

    def g(row, col):
        i = idx.get(col)
        return row[i] if i is not None and i < len(row) else None

    rows, skipped = [], 0
    for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
        d, code = to_date(g(row, "날짜")), norm_code(g(row, "종목코드"))
        if not d or not code:
            skipped += 1
            continue
        rows.append((
            d, code,
            to_num(g(row, "시가")), to_num(g(row, "고가")), to_num(g(row, "저가")),
            to_num(g(row, "종가")), to_num(g(row, "대비")), to_num(g(row, "등락률")),
            to_num(g(row, "거래량")), to_num(g(row, "거래대금")),
            to_num(g(row, "시가총액")), to_num(g(row, "상장주식수")),
            "EXCEL",
        ))
    wb.close()
    print(f"  파싱 {len(rows):,}행 (스킵 {skipped:,})")
    rows = dedup(rows, 2, "전종목시세")

    with conn() as c, c.cursor() as cur:
        # FK 위반 방지 — 마스터에 없는 종목 선등록
        cur.execute("select code from stocks")
        known = {r[0] for r in cur.fetchall()}
        missing = {r[1] for r in rows} - known
        if missing:
            print(f"  마스터에 없던 종목 {len(missing):,}개 선등록")
            execute_values(cur,
                "insert into stocks(code,name) values %s on conflict do nothing",
                [(m, m) for m in missing])

        push(cur, """
            insert into daily_price
              (trade_date,code,open,high,low,close,change,change_pct,
               volume,trade_amount,market_cap,listed_shares,source)
            values %s
            on conflict (trade_date,code) do update set
              open=excluded.open, high=excluded.high, low=excluded.low,
              close=excluded.close, change=excluded.change, change_pct=excluded.change_pct,
              volume=excluded.volume, trade_amount=excluded.trade_amount,
              market_cap=excluded.market_cap, listed_shares=excluded.listed_shares,
              source=excluded.source
        """, rows, "적재")
        log(cur, "price", "SUCCESS", len(rows), "from 전종목시세.xlsx")
        c.commit()
    print("  ✅ 완료 (weight_per_share는 DB가 자동 계산)")


# ──────────────────────────────────────────────────────────────
# ③ 수급주체정리 → daily_flow  (7개 시트를 (날짜,종목)으로 병합)
# ──────────────────────────────────────────────────────────────
def migrate_flow():
    print("\n[3/4] 수급주체정리 → daily_flow")
    need = ["날짜", "종목코드", "거래대금_순매수"]
    merged = {}          # (date, code) -> {컬럼: 값}
    partial = set()      # 필터 수집분 (date, code)

    wb = openpyxl.load_workbook(XL_FLOW, read_only=True, data_only=True)
    for sheet, col in FLOW_SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"  ⚠ '{sheet}' 시트 없음")
            continue
        ws = wb[sheet]
        hdr, idx = find_header(ws, need)
        if not idx:
            print(f"  ⚠ '{sheet}' 헤더 못 찾음")
            continue

        i_d, i_c, i_n = idx["날짜"], idx["종목코드"], idx["거래대금_순매수"]
        i_v = idx.get("거래량_순매수")
        n = 0
        for row in ws.iter_rows(min_row=hdr + 1, values_only=True):
            if i_d >= len(row) or i_c >= len(row):
                continue
            d, code = to_date(row[i_d]), norm_code(row[i_c])
            if not d or not code:
                continue
            e = merged.setdefault((d, code), {})
            e[col] = to_num(row[i_n]) if i_n < len(row) else None
            if i_v is not None and i_v < len(row):
                if col == "foreign_net":
                    e["foreign_net_vol"] = to_num(row[i_v])
                elif col == "inst_net":
                    e["inst_net_vol"] = to_num(row[i_v])
            if sheet in PARTIAL_SHEETS:
                partial.add((d, code))
            n += 1
        print(f"  {sheet:<9} {n:>7,}행 (헤더 {hdr}행)")
    wb.close()

    # corp_other_net·foreign_net_vol·inst_net_vol은 2026-08 용량 정리 때 daily_flow에서
    # 컬럼 삭제 (신호 엔진 미사용). 파싱 로직(merged 딕셔너리)은 그대로 두고 여기서만 제외.
    cols = ["foreign_net", "inst_net", "fin_inv_net", "inv_trust_net",
            "pension_net", "pe_net"]
    rows = [
        (d, code, *[e.get(c) for c in cols], "EXCEL", (d, code) in partial)
        for (d, code), e in merged.items()
    ]
    print(f"  → 병합 {len(rows):,}행 (is_partial {len(partial):,}행)")

    with conn() as c, c.cursor() as cur:
        cur.execute("select code from stocks")
        known = {r[0] for r in cur.fetchall()}
        missing = {r[1] for r in rows} - known
        if missing:
            print(f"  마스터에 없던 종목 {len(missing):,}개 선등록")
            execute_values(cur,
                "insert into stocks(code,name) values %s on conflict do nothing",
                [(m, m) for m in missing])

        push(cur, f"""
            insert into daily_flow (trade_date,code,{','.join(cols)},source,is_partial)
            values %s
            on conflict (trade_date,code) do update set
              {', '.join(f'{c}=coalesce(excluded.{c}, daily_flow.{c})' for c in cols)},
              source=excluded.source,
              is_partial=excluded.is_partial
        """, rows, "적재")
        log(cur, "flow", "SUCCESS", len(rows), f"partial={len(partial)}")
        c.commit()
    print("  ✅ 완료")


# ──────────────────────────────────────────────────────────────
# ④ 키움 보유종목 통계 CSV
# ──────────────────────────────────────────────────────────────
def migrate_kiwoom(path=None):
    path = path or CSV_KIWOOM
    print(f"\n[4/4] 키움 보유종목 통계 → kiwoom_holder_stats\n  파일 {path}")

    # 파일명 앞 8자리(YYYYMMDD)를 기준일로 사용
    m = re.search(r"(20\d{6})", os.path.basename(path))
    if not m:
        sys.exit("❌ 파일명에서 날짜(YYYYMMDD)를 찾지 못했습니다. 예: 20260814_키움보유종목통계.csv")
    d = f"{m.group(1)[:4]}-{m.group(1)[4:6]}-{m.group(1)[6:]}"
    print(f"  기준일 {d}")

    rows = []
    with open(path, encoding="cp949") as f:
        for r in csv.DictReader(f):
            code = norm_code(r.get("종목코드"))
            if not code:
                continue
            rows.append((
                d, code,
                to_num(r.get("순위")),
                to_num(r.get("보유계좌수")),
                to_num(r.get("평균매수가")),
                to_num(r.get("보유계좌수익률")),
                to_num(r.get("현재가")),
            ))
    print(f"  파싱 {len(rows):,}행")
    rows = dedup(rows, 2, "키움 CSV")

    with conn() as c, c.cursor() as cur:
        push(cur, """
            insert into kiwoom_holder_stats
              (trade_date,code,rank,accounts,avg_buy_price,return_pct,close_at_upload)
            values %s
            on conflict (trade_date,code) do update set
              rank=excluded.rank, accounts=excluded.accounts,
              avg_buy_price=excluded.avg_buy_price, return_pct=excluded.return_pct,
              close_at_upload=excluded.close_at_upload, uploaded_at=now()
        """, rows, "적재")
        log(cur, "kiwoom", "SUCCESS", len(rows), d)
        c.commit()
    print("  ✅ 완료")


# ──────────────────────────────────────────────────────────────
# 점검
# ──────────────────────────────────────────────────────────────
def check():
    print("\n적재 결과 점검")
    with conn() as c, c.cursor() as cur:
        for q, label in [
            ("select count(*) from stocks", "stocks"),
            ("select count(*) from stocks where security_type='STOCK'", "  └ STOCK"),
            ("select count(*) from stocks where sector_krx is not null", "  └ 업종 있음"),
            ("select count(*) from daily_price", "daily_price"),
            ("select count(*) from daily_flow", "daily_flow"),
            ("select count(*) from daily_flow where is_partial", "  └ is_partial"),
            ("select count(*) from kiwoom_holder_stats", "kiwoom"),
        ]:
            cur.execute(q)
            print(f"  {label:<16} {cur.fetchone()[0]:>10,}")

        cur.execute("select min(trade_date),max(trade_date),count(distinct trade_date) from daily_price")
        a, b, n = cur.fetchone()
        print(f"\n  시세 기간  {a} ~ {b}  ({n}거래일)")
        cur.execute("select min(trade_date),max(trade_date),count(distinct trade_date) from daily_flow")
        a, b, n = cur.fetchone()
        print(f"  수급 기간  {a} ~ {b}  ({n}거래일)")

        print("\n  파생지표 검산 (엑셀 무게/주식수와 대조):")
        # amt_cap_ratio는 2026-08 용량 정리 때 컬럼 삭제 — weight_per_share만 검산.
        cur.execute("""
            select code, round(weight_per_share,10)
            from daily_price
            where trade_date='2025-09-01' and code in ('282330','138930')
            order by code
        """)
        for code, w in cur.fetchall():
            print(f"    {code}  무게/주식수 {w}")
        print("    (엑셀: 138930 → 0.0028872133,  282330 → -0.0055692018)")

        print("\n  최근 5거래일 적재 현황:")
        cur.execute("select * from v_data_coverage limit 5")
        print(f"    {'날짜':<12}{'시세':>8}{'수급':>8}{'키움':>8}{'partial':>9}")
        for r in cur.fetchall():
            print(f"    {str(r[0]):<12}{r[1]:>8,}{r[2]:>8,}{r[3]:>8,}{r[4]:>9,}")


if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "check"
    t0 = time.time()
    if cmd == "stocks":   migrate_stocks()
    elif cmd == "price":  migrate_price()
    elif cmd == "flow":   migrate_flow()
    elif cmd == "kiwoom": migrate_kiwoom(sys.argv[2] if len(sys.argv) > 2 else None)
    elif cmd == "all":
        migrate_stocks(); migrate_price(); migrate_flow(); migrate_kiwoom(); check()
    elif cmd == "check":  check()
    else:
        sys.exit(f"알 수 없는 명령: {cmd}\n사용: stocks | price | flow | kiwoom | all | check")
    print(f"\n소요 {time.time()-t0:.0f}초")
